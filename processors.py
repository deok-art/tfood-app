from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, timedelta
from io import BytesIO, StringIO
from typing import Callable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


UNKNOWN_VENDOR = "미인식"
LOTTE_MART_VENDOR_NAME = "롯데쇼핑㈜ 롯데마트사업본부"

# 이력번호 prefix + 이력사이트 제품명
# key: 상품명에 포함된 키워드
PRODUCT_CODE_MAP: dict[str, tuple[str, str]] = {
    "가리비치즈진밥":    ("11586725", "베이비본죽 가리비치즈진밥 180g"),
    "게살보리진밥":      ("11586625", "베이비본죽 게살보리진밥 180g"),
    "닭고기버섯죽":      ("6572719",  "베이비본죽 닭고기버섯죽"),
    "닭고기알밤진밥":    ("9916423",  "베이비본죽 닭고기알밤진밥"),
    "닭고기애호박미역죽": ("7506320", "베이비본죽 닭고기애호박미역죽"),
    "닭고기양송이진밥":  ("6572819",  "베이비본죽 닭고기양송이진밥"),
    "오트밀버섯전복죽":  ("10878124", "베이비본죽 오트밀버섯전복죽 180"),
    "전복영양진밥":      ("10878524", "베이비본죽 전복영양진밥 180"),
    "찹쌀누룽지닭죽":    ("9916723",  "베이비본죽 찹쌀누룽지닭죽"),
    "퀴노아미역전복죽":  ("11586525", "베이비본죽 퀴노아미역전복죽 180g"),
    "한우과일죽":        ("7506220",  "베이비본죽 한우과일죽"),
    "한우버섯무죽":      ("10878324", "베이비본죽 한우버섯무죽 180"),
    "한우불고기진밥":    ("9916223",  "베이비본죽 한우불고기진밥"),
    "한우뿌리채소죽":    ("9916523",  "베이비본죽 한우뿌리채소죽"),
    "한우사골진밥":      ("10878624", "베이비본죽 한우사골진밥 180"),
    "한우야채진밥":      ("6572419",  "베이비본죽 한우야채진밥"),
    "한우참깨애호박죽":  ("6572519",  "베이비본죽 한우참깨애호박죽"),
    "흰살생선채소죽":    ("11586425", "베이비본죽 흰살생선채소죽 180g"),
    # 신규 제품 — 식품이력추적관리번호 prefix 미확인
    "한우치즈영양밥":   ("",         "베이비본죽 한우치즈영양밥"),
    "전복버터영양밥":   ("",         "베이비본죽 전복버터영양밥"),
    "미트카레영양밥":   ("",         "베이비본죽 미트카레영양밥"),
    "닭살들깨버섯영양밥": ("",       "베이비본죽 닭살들깨버섯영양밥"),
}

# 이지어드민 출고처 → 이력사이트 출고처명
CHANNEL_NAME_MAP: dict[str, str] = {
    "쿠팡":  "쿠팡(주)",
    "컬리":  "주식회사 컬리",
    "개인":  "개인",
    "네이버": "주식회사 네이버",
}

OUTPUT_COLUMNS = [
    "전송여부", "출고일", "정보연계일자",
    "식품이력추적관리번호", "제품명", "출고처명", "출고수량",
]

# 이지어드민 중간 검토 형식 컬럼
REVIEW_COLUMNS = ["작업일자", "소비기한", "상품명", "출고처", "작업수량", "비고"]

# 이지어드민 짧은 채널명 → 이력사이트 공식 거래처명 (이력사이트 등록 시 사용)
CHANNEL_FULL_NAME: dict[str, str] = {
    "컬리":  "주식회사 컬리",
    "쿠팡":  "쿠팡(주)",
    "네이버": "주식회사 네이버",
    "개인":  "개인",
}

VENDOR_ALIASES: dict[str, list[str]] = {
    "이지어드민": ["작업일자", "작업수량", "메모"],
    "롯데마트":   ["출고예정일자", "출고수량(개수)"],
    "명현유통":   ["명현유통"],
    "본에프디":   ["본에프디", "togo"],
}


@dataclass
class InputRecord:
    source_type: str
    source_name: str
    dataframe: pd.DataFrame
    first_row_text: str
    detected_vendor: str
    selected_vendor: str


Processor = Callable[[pd.DataFrame], pd.DataFrame]


def _match_product(name: str) -> tuple[str, str] | None:
    cleaned_name = str(name).replace(" ", "")
    for keyword, info in PRODUCT_CODE_MAP.items():
        if keyword.replace(" ", "") in cleaned_name:
            return info
    return None


def _to_date(value: object) -> date | None:
    try:
        return pd.to_datetime(str(value)).date()
    except Exception:
        return None


def _tracking_number(prefix: str, expiry_raw: str) -> str:
    return prefix + expiry_raw.replace("-", "").strip()


def _pass_through(df: pd.DataFrame) -> pd.DataFrame:
    return df.copy()


def _find_col(df: pd.DataFrame, *keywords: str) -> str | None:
    for kw in keywords:
        for col in df.columns:
            if kw in str(col):
                return col
    return None


# 메모 → 출고처 규칙. 위에서부터 첫 매칭 적용.
# 새 패턴 학습되면 여기 한 줄씩 추가. (키워드, 출고처)
CHANNEL_MEMO_RULES: list[tuple[str, str]] = [
    ("쿠팡출하", "쿠팡"),
    ("컬리출하", "컬리"),
    ("네이버출하", "네이버"),
    ("ezpos", "개인"),
    ("송장출력", "개인"),
]

# 어떤 규칙에도 안 걸린 메모 → 휴먼터치 필요
NEEDS_REVIEW_CHANNEL = "❓확인필요"


def _parse_channel_from_memo(memo: str) -> str:
    """메모에서 출고처 판정. 규칙 미매칭 시 '확인필요'로 띄움(개인으로 흡수 안 함)."""
    m = memo.strip()
    for keyword, channel in CHANNEL_MEMO_RULES:
        if keyword in m:
            return channel
    return NEEDS_REVIEW_CHANNEL


def _process_ezadmin(df: pd.DataFrame) -> pd.DataFrame:
    """이지어드민 출고 raw → 중간 검토 형식 (REVIEW_COLUMNS).

    실제 컬럼 구조 (17개):
      작업일자 / 입고일 / 제조일 / 유통기한 / 로트번호 / 제조번호 /
      로케이션 / 상품코드 / 상품명 / 옵션 / 이전재고 / 이후재고 /
      작업수량 / 작업 / 현재고(평치+일반) / 메모 / 바코드

    출고처는 '메모' 컬럼에 인코딩되어 있음.
    집계: (소비기한, 상품명, 출고처) 동일 조합의 수량 합산.
    """
    date_col    = _find_col(df, "작업일자")
    expiry_col  = _find_col(df, "유통기한")
    product_col = _find_col(df, "상품명")
    # "작업" 컬럼 exact match — 작업일자·작업수량과 구분
    type_col    = next((c for c in df.columns if str(c).strip() == "작업"), None)
    memo_col    = _find_col(df, "메모")
    qty_col     = _find_col(df, "작업수량")

    missing = [name for name, col in [
        ("작업일자", date_col), ("유통기한", expiry_col),
        ("상품명", product_col), ("작업", type_col),
        ("메모", memo_col), ("작업수량", qty_col),
    ] if col is None]
    if missing:
        raise ValueError(f"이지어드민 필수 컬럼 없음: {missing}")

    # 입고 제외 → 배송·출고만 처리
    df = df[df[type_col].astype(str).isin(["배송", "출고"])].copy()
    # '직납', '영양밥' 상품 제외
    df = df[~df[product_col].astype(str).str.contains("직납|영양밥", na=False, regex=True)].copy()

    rows: list[dict] = []
    for _, row in df.iterrows():
        try:
            dt = pd.to_datetime(str(row[date_col]).strip())
        except Exception:
            continue

        product_name = str(row[product_col]).strip()
        expiry_raw   = str(row[expiry_col]).strip()  # YYYY-MM-DD
        memo         = str(row[memo_col]).strip()
        try:
            qty_str = str(row[qty_col]).replace(",", "").strip()
            qty = int(float(qty_str))
        except (ValueError, TypeError):
            continue
        if qty <= 0:
            continue

        rows.append({
            "작업일자": dt,
            "소비기한": expiry_raw,
            "상품명": product_name,
            "출고처": _parse_channel_from_memo(memo),
            "작업수량": qty,
            "비고": memo,
        })

    if not rows:
        return pd.DataFrame(columns=REVIEW_COLUMNS)

    result = pd.DataFrame(rows)
    result["작업수량"] = pd.to_numeric(result["작업수량"], errors="coerce").fillna(0).astype(int)

    # (소비기한, 상품명, 출고처) 동일 조합 합산 — 복수 박스 스캔 집계
    agg = (
        result
        .groupby(["소비기한", "상품명", "출고처"], sort=False, as_index=False)
        .agg(작업일자=("작업일자", "max"), 작업수량=("작업수량", "sum"), 비고=("비고", lambda x: " | ".join(sorted(set(x)))))
    )
    # 확인필요가 아닌 행은 비고란을 비워 깔끔하게 표시
    agg["비고"] = agg.apply(lambda row: row["비고"] if row["출고처"] == NEEDS_REVIEW_CHANNEL else "", axis=1)

    # 작업일자 HH:MM 표시 (초 생략)
    agg["작업일자"] = agg["작업일자"].dt.strftime("%Y-%m-%d %H:%M")

    agg = agg.sort_values(["상품명", "소비기한", "출고처"]).reset_index(drop=True)
    return agg[REVIEW_COLUMNS]


def _process_lottemart(df: pd.DataFrame) -> pd.DataFrame:
    """롯데마트 출고 raw → 이력사이트 출고 등록 형식."""
    date_col    = _find_col(df, "출고예정일자", "출고일자")
    product_col = _find_col(df, "품목명")
    expiry_col  = _find_col(df, "소비기한")
    qty_col     = _find_col(df, "출고수량(개수)", "개수")

    missing = [k for k, c in [("출고예정일자", date_col), ("품목명", product_col),
                               ("소비기한", expiry_col), ("출고수량(개수)", qty_col)] if c is None]
    if missing:
        raise ValueError(f"롯데마트 필수 컬럼 없음: {missing}")

    rows: list[dict] = []
    for _, row in df.iterrows():
        ship_date = _to_date(row[date_col])
        if ship_date is None:
            continue

        product_name = str(row[product_col])
        if "영양밥" in product_name:
            continue
        expiry_raw   = str(row[expiry_col]).strip()  # YYYYMMDD 형식 (대시 없음)
        qty          = row[qty_col]

        link_date = ship_date + timedelta(days=7)

        match = _match_product(product_name)
        if match is None:
            rows.append({
                "전송여부": "⚠미인식",
                "출고일": ship_date.isoformat(),
                "정보연계일자": link_date.isoformat(),
                "식품이력추적관리번호": f"[미인식] {product_name}",
                "제품명": product_name,
                "출고처명": LOTTE_MART_VENDOR_NAME,
                "출고수량": qty,
            })
            continue

        prefix, official_name = match
        rows.append({
            "전송여부": "전송",
            "출고일": ship_date.isoformat(),
            "정보연계일자": link_date.isoformat(),
            "식품이력추적관리번호": _tracking_number(prefix, expiry_raw),
            "제품명": official_name,
            "출고처명": LOTTE_MART_VENDOR_NAME,
            "출고수량": qty,
        })

    return pd.DataFrame(rows, columns=OUTPUT_COLUMNS) if rows else pd.DataFrame(columns=OUTPUT_COLUMNS)


PROCESSOR_REGISTRY: dict[str, Processor] = {
    UNKNOWN_VENDOR: _pass_through,
    "이지어드민": _process_ezadmin,
    "롯데마트":   _process_lottemart,
    "명현유통":   _pass_through,
    "본에프디":   _pass_through,
}


def available_vendors() -> list[str]:
    vendors = sorted(set(VENDOR_ALIASES) | set(PROCESSOR_REGISTRY))
    if UNKNOWN_VENDOR not in vendors:
        vendors.insert(0, UNKNOWN_VENDOR)
    return vendors


def first_row_as_text(dataframe: pd.DataFrame) -> str:
    if dataframe.empty:
        return ""
    first_row = dataframe.iloc[0].fillna("").astype(str).tolist()
    return " ".join(v.strip() for v in first_row if v.strip())


def detect_vendor(source_name: str, dataframe: pd.DataFrame) -> str:
    # 컬럼 헤더로 우선 감지
    columns_text = " ".join(dataframe.columns.astype(str))
    haystack = f"{source_name} {columns_text} {first_row_as_text(dataframe)}".casefold()

    best_vendor = UNKNOWN_VENDOR
    best_score = 0
    for vendor_id, aliases in VENDOR_ALIASES.items():
        score = sum(1 for a in aliases if a.casefold() in haystack)
        if score > best_score:
            best_score = score
            best_vendor = vendor_id

    return best_vendor if best_score >= 1 else UNKNOWN_VENDOR


def read_excel_upload(uploaded_file) -> pd.DataFrame:
    return pd.read_excel(BytesIO(uploaded_file.getvalue()), dtype=object)


def read_clipboard_table(raw_text: str) -> pd.DataFrame:
    cleaned = raw_text.strip()
    if not cleaned:
        return pd.DataFrame()
    return pd.read_csv(StringIO(cleaned), sep="\t", dtype=object)


def build_record(source_type: str, source_name: str, dataframe: pd.DataFrame) -> InputRecord:
    detected_vendor = detect_vendor(source_name, dataframe)
    return InputRecord(
        source_type=source_type,
        source_name=source_name,
        dataframe=dataframe,
        first_row_text=first_row_as_text(dataframe),
        detected_vendor=detected_vendor,
        selected_vendor=detected_vendor,
    )


def process_record(record: InputRecord) -> pd.DataFrame:
    processor = PROCESSOR_REGISTRY.get(record.selected_vendor, _pass_through)
    processed = processor(record.dataframe)
    if record.selected_vendor in ("이지어드민", "롯데마트"):
        return processed
    # 미인식/기타 거래처는 원본 컬럼 유지 + 거래처 컬럼 추가
    processed = processed.copy()
    processed.insert(0, "거래처", record.selected_vendor)
    processed.insert(1, "입력출처", record.source_type)
    return processed


def combine_records(records: list[InputRecord]) -> pd.DataFrame:
    frames = [process_record(r) for r in records]
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


# ── 엑셀 스타일 팔레트 (RRGGBB) ──────────────────────────────────────
_HEADER_FILL = "1F4E2C"   # 딥그린 — 헤더·합계 배경
_HEADER_FONT = "FFFFFF"   # 흰색 글씨
_GROUP_FILL_A = "FFFFFF"  # 상품 그룹 zebra — 밝은 행
_GROUP_FILL_B = "EDF4EF"  # 상품 그룹 zebra — 연그린 행
_BORDER_COLOR = "D9D9D9"  # 얇은 회색 테두리
_EXPIRY_FILL = "FFF2CC"   # 연주황 — 소비기한 다중 상품 강조 배경
_EXPIRY_FONT = "9C6500"   # 진한 주황 — 소비기한 다중 상품 글씨
_EXPIRY_BORDER = "BF8F00"  # 골드 — 같은 상품 내 소비기한 경계 구분선

# 출고처별 (배경, 글씨) — 채널 브랜드 톤
_CHANNEL_STYLE: dict[str, tuple[str, str]] = {
    "개인":  ("ECECEC", "595959"),  # 회색
    "컬리":  ("E7DCF2", "7030A0"),  # 컬리 퍼플
    "쿠팡":  ("FCE0E0", "C00000"),  # 쿠팡 레드
    "네이버": ("DCEFE0", "1E7B34"),  # 네이버 그린
    NEEDS_REVIEW_CHANNEL: ("FFC7CE", "9C0006"),  # 빨강 — 휴먼터치 필요
}

# 컬럼별 너비·정렬 (컬럼명: (너비, 가운데정렬여부))
_COLUMN_LAYOUT: dict[str, tuple[int, bool]] = {
    "작업일자": (18, True),
    "소비기한": (13, True),
    "상품명":   (30, False),
    "출고처":   (10, True),
    "작업수량": (11, True),
    "비고":    (40, False),
}


def _thin_border() -> Border:
    side = Side(style="thin", color=_BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)


def _styled_review_excel(df: pd.DataFrame) -> bytes:
    """이지어드민 중간 검토 형식(REVIEW_COLUMNS)을 색상 구분된 엑셀로 변환.

    - 헤더: 딥그린 배경 + 흰 볼드, 틀 고정
    - 출고처: 채널별 고유 색
    - 상품명: 같은 상품 그룹끼리 교대 음영(zebra)
    - 합계 행: 딥그린 강조
    """
    columns = list(df.columns)
    border = _thin_border()
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "이지어드민 출고검토"

    # ── 헤더 행 ──
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    header_font = Font(color=_HEADER_FONT, bold=True, size=11)
    ws.append(columns)
    for col_idx, name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[1].height = 24

    # ── 데이터 행 (상품 그룹 zebra + 소비기한 구분) ──
    qty_idx = columns.index("작업수량") + 1 if "작업수량" in columns else None
    product_idx = columns.index("상품명") if "상품명" in columns else None
    channel_idx = columns.index("출고처") if "출고처" in columns else None
    expiry_idx = columns.index("소비기한") if "소비기한" in columns else None

    # 상품별 고유 소비기한 개수 → 2개 이상이면 소비기한 셀 강조
    expiry_count: dict[str, int] = {}
    if product_idx is not None and expiry_idx is not None:
        expiry_count = (
            df.groupby(columns[product_idx])[columns[expiry_idx]].nunique().to_dict()
        )
    boundary_side = Side(style="medium", color=_EXPIRY_BORDER)
    expiry_fill = PatternFill("solid", fgColor=_EXPIRY_FILL)
    expiry_font = Font(color=_EXPIRY_FONT, bold=True)

    prev_product: str | None = None
    prev_expiry: str | None = None
    group_toggle = False
    excel_row = 2
    for _, record in df.iterrows():
        values = [record[c] for c in columns]
        current_product = str(values[product_idx]) if product_idx is not None else None
        current_expiry = str(values[expiry_idx]) if expiry_idx is not None else None

        # 상품명이 바뀌면 음영 토글 → 같은 상품 그룹은 동일 배경
        is_new_product = current_product != prev_product
        if is_new_product:
            group_toggle = not group_toggle
        # 같은 상품인데 소비기한만 바뀌는 경계 행 → 윗변 굵은 구분선
        is_expiry_boundary = (
            not is_new_product
            and current_expiry is not None
            and current_expiry != prev_expiry
        )

        group_fill = PatternFill(
            "solid", fgColor=_GROUP_FILL_B if group_toggle else _GROUP_FILL_A
        )
        row_border = border
        if is_expiry_boundary:
            row_border = Border(
                left=border.left, right=border.right,
                top=boundary_side, bottom=border.bottom,
            )

        ws.append(values)
        for col_idx, name in enumerate(columns, 1):
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.fill = group_fill
            cell.border = row_border
            cell.alignment = left if name == "상품명" else center

        # 소비기한이 여러 개인 상품: 소비기한 셀 강조
        if expiry_idx is not None and expiry_count.get(current_product, 0) >= 2:
            ec = ws.cell(row=excel_row, column=expiry_idx + 1)
            ec.fill = expiry_fill
            ec.font = expiry_font

        # 출고처 셀: 채널별 색 덮어쓰기
        if channel_idx is not None:
            ch = str(values[channel_idx])
            if ch in _CHANNEL_STYLE:
                bg, fg = _CHANNEL_STYLE[ch]
                ch_cell = ws.cell(row=excel_row, column=channel_idx + 1)
                ch_cell.fill = PatternFill("solid", fgColor=bg)
                ch_cell.font = Font(color=fg, bold=True)

        # 작업수량: 천단위 콤마
        if qty_idx is not None:
            ws.cell(row=excel_row, column=qty_idx).number_format = "#,##0"

        ws.row_dimensions[excel_row].height = 18
        prev_product = current_product
        prev_expiry = current_expiry
        excel_row += 1

    # ── 합계 행 ──
    total_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    total_font = Font(color=_HEADER_FONT, bold=True)
    total_values = ["" for _ in columns]
    if columns:
        total_values[0] = "합계"
    if qty_idx is not None:
        col_letter = get_column_letter(qty_idx)
        if excel_row > 2:
            total_values[qty_idx - 1] = f"=SUM({col_letter}2:{col_letter}{excel_row-1})"
        else:
            total_values[qty_idx - 1] = 0
    ws.append(total_values)
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=excel_row, column=col_idx)
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = center
        cell.border = border
    if qty_idx is not None:
        ws.cell(row=excel_row, column=qty_idx).number_format = "#,##0"
    ws.row_dimensions[excel_row].height = 22

    # ── 열너비 · 틀 고정 ──
    for col_idx, name in enumerate(columns, 1):
        width, _ = _COLUMN_LAYOUT.get(name, (14, True))
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _extract_mmdd(dataframe: pd.DataFrame) -> str:
    """데이터의 작업일자/출고일에서 MMDD 추출. 실패 시 오늘 날짜."""
    for col in ("작업일자", "출고일"):
        if col in dataframe.columns and not dataframe.empty:
            try:
                return pd.to_datetime(str(dataframe[col].iloc[0])).strftime("%m%d")
            except Exception:
                continue
    return date.today().strftime("%m%d")


def build_export_filename(dataframe: pd.DataFrame) -> str:
    """다운로드 파일명 생성. 이지어드민 검토 형식이면 '이지어드민 출고현황 MMDD.xlsx'."""
    mmdd = _extract_mmdd(dataframe)
    if list(dataframe.columns) == REVIEW_COLUMNS:
        return f"이지어드민 출고현황 {mmdd}.xlsx"
    return f"출고취합 {mmdd}.xlsx"


def dataframe_to_excel_bytes(dataframe: pd.DataFrame) -> bytes:
    # 이지어드민 중간 검토 형식이면 색상 구분 엑셀로 출력
    if list(dataframe.columns) == REVIEW_COLUMNS:
        return _styled_review_excel(dataframe)

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name="출고이력등록")
    buffer.seek(0)
    return buffer.getvalue()
